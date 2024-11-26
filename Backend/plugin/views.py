from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from pathlib import Path
from .models import * 
import requests
from rest_framework.parsers import MultiPartParser
import json
from .view.register import register
from .view.verify_license_id  import verify_lid
from .services.check_email import check_email

from users.models import PluginMaster, License,LicenseAllocation
from django.shortcuts import render
from django.contrib.auth.tokens import default_token_generator
from rest_framework import viewsets, permissions,generics,status,mixins
from rest_framework.decorators import action
from .serializers import * 
from plugin.models import * 
from rest_framework.response import Response 
from django.contrib.auth import get_user_model, authenticate
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views.decorators.http import require_http_methods
from plugin.serializers import  *
from django.db import connection
import logging
from django.core.files.storage import default_storage
import os
from django.core.files.base import ContentFile
from django.core.serializers import serialize
logger = logging.getLogger(__name__)
from users.serializers import *
from openpyxl import load_workbook 
import re
from PyPDF2 import PdfReader, PdfWriter
from rest_framework.views import APIView
import time
from django.utils.timezone import now
from rest_framework.decorators import api_view
@csrf_exempt
def registration_view(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        
        # Extract required fields from input data
        license_id = data.get('licenseId')
        plugin_id = data.get('pluginId')
        ip_add = data.get('ipAddress')
        browser = data.get('browser')
        uuid = data.get('uuid')
        mac_address = data.get('macAddress')
        
        # Validate mandatory fields
        if not license_id or not uuid or not mac_address or not browser:
            return JsonResponse({
                "message": "Missing required fields (licenseId, uuid, macAddress, browser)",
                "STATUS": "Error",
                "Code": 0
            }, status=400)

        try:
            license = License.objects.get(hashed_license_id=license_id)
            allocation = LicenseAllocation.objects.filter(license=license).order_by('-allocation_date').first()
            
            if not allocation:
                return JsonResponse({
                    "message": "License found but no allocation. Please contact the administrator.",
                    "STATUS": "Allocation Not Found",
                    "Code": 0
                }, status=400)

            existing_system = UserSystemDetails.objects.filter(license_allocation=allocation, uuid=uuid).first()
            
            if existing_system:
                if existing_system.mac_address == mac_address:
                    # Check if browser exists and is active
                    browser_details = SystemBrowserDetails.objects.filter(
                        device_information=existing_system,
                        browser=browser
                    ).first()
                    
                    if browser_details:
                        if browser_details.is_active:
                            return JsonResponse({
                                "message": "This device with this browser is already registered and active.",
                                "STATUS": "Already Registered",
                                "Code": 0
                            }, status=400)
                        else:
                            # Reactivate the browser
                            browser_details.is_active = True
                            browser_details.unregistered_at = None
                            browser_details.ipv4 = ip_add
                            browser_details.save()
                            
                            return JsonResponse({
                                "message": "Browser reactivated successfully.",
                                "STATUS": "Browser Reactivated",
                                "Code": 1,
                                "data": {"email": allocation.allocated_to}
                            }, status=200)
                    else:
                        # Register new browser
                        SystemBrowserDetails.objects.create(
                            device_information=existing_system,
                            ipv4=ip_add,
                            browser=browser,
                            is_active=True
                        )
                        return JsonResponse({
                            "message": "New browser registered successfully.",
                            "STATUS": "Browser Registered",
                            "Code": 1,
                            "data": {"email": allocation.allocated_to}
                        }, status=200)
                else:
                    return JsonResponse({
                        "message": "License is already registered on another device.",
                        "STATUS": "Device Mismatch",
                        "Code": 0
                    }, status=400)

            # Create new system and browser registration
            with transaction.atomic():
                user_system_details, created = UserSystemDetails.objects.update_or_create(
                    uuid=uuid,
                    defaults={
                        "license_allocation": allocation,
                        "mac_address": mac_address,
                        "serial_number": data.get('serialNumber'),
                        "os_type": data.get('osType'),
                        "os_platform": data.get('osPlatform'),
                        "os_release": data.get('osRelease'),
                        "host_name": data.get('hostName'),
                        "architecture": data.get('architecture'),
                    }
                )
                
                SystemBrowserDetails.objects.create(
                    device_information=user_system_details,
                    ipv4=ip_add,
                    browser=browser,
                    is_active=True
                )

                plugin_install_uninstall = PluginInstallUninstall.objects.create(
                    plugin_id=plugin_id,
                    ip_address=ip_add,
                    browser=browser,
                    installed_at=timezone.now()
                )

                plugin, _ = PluginMaster.objects.update_or_create(
                    plugin_id=plugin_id,
                    defaults={
                        'license_id': license,
                        'ip_add': ip_add,
                        'browser': browser,
                        'install_date': timezone.now()
                    }
                )

                enable_disable_action = PluginEnableDisable.objects.create(
                    plugin_install_uninstall=plugin_install_uninstall,
                    enabled_at=timezone.now()
                )

                return verify_lid(data)

        except License.DoesNotExist:
            return JsonResponse({
                "message": "License not found",
                "STATUS": "Not Found",
                "Code": 0
            }, status=404)
        
    return JsonResponse({
        "error": "Invalid request method",
        "Code": 0
    }, status=405)

@csrf_exempt
def verify_license_id_view(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        response = verify_lid(data)
        return response
        
@csrf_exempt
def check_email_view(request):
    if request.method == 'POST':
        try:
            return check_email(request)
        except Exception as e:
            logger.error(f"Error in check_email_view: {e}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({"error": "Invalid request method"}, status=405)
        
@csrf_exempt
def spam_email_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            response = spam_email(data)
            return response
        except Exception as e:
            
            return str(e)
        
        
    return JsonResponse({"error": "Invalid request method"}, status=405)


class DisputeViewSet(viewsets.ViewSet):
    serializer_class = DisputeSerializer
    #permission_classes = [IsAuthenticated]


    @action(detail=False, methods=['get'], url_path='count')
    def get_dispute_count(self, request):
        email = request.query_params.get('email')
        msg_id = request.query_params.get('messageId')

        if not email:
            return Response({"error": "Email is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        if not msg_id:
            return Response({"error": "Message Id is required"}, status=status.HTTP_400_BAD_REQUEST)

        
        active_dispute_count = Dispute.objects.filter(email=email, msg_id=msg_id).count()

        try:
            email_detail = EmailDetails.objects.get(message_id=msg_id)
        except EmailDetails.DoesNotExist:
            return Response({"error": "Message Id not found in email_details"}, status=status.HTTP_404_NOT_FOUND)

        return Response({
            "dispute_count": active_dispute_count,
            "message_status": email_detail.status,
            "message_id": msg_id,
            "subject": email_detail.subject,
            "plugin_id": email_detail.plugin_id
        }, status=status.HTTP_200_OK)
        

class PluginInstallUninstallViewSet(viewsets.ViewSet):
    """
    A ViewSet for handling plugin installation and uninstallation actions.
    """

    @action(detail=False, methods=['post'], url_path='install')
    def install(self, request):
        # plugin id is send from frontend 
        # ip address is send from frontend 
        plugin_id = request.data.get('pluginId')
        ip_address = request.data.get('ipAddress')

        if not all([plugin_id, ip_address]):
            return Response({'error': 'All fields (plugin_id,  ip_address) are required'}, status=status.HTTP_400_BAD_REQUEST)

    # Check if a plugin with the same plugin_id exists and is not uninstalled
        existing_plugin = PluginInstallUninstall.objects.filter(plugin_id=plugin_id, uninstalled_at__isnull=True).first()
    
        if existing_plugin:
            return Response({'error': 'The plugin is already installed and not yet uninstalled.'}, status=status.HTTP_400_BAD_REQUEST)

    # Create a new entry for plugin installation if no active installation exists
        plugin_action = PluginInstallUninstall.objects.create(
            plugin_id=plugin_id,
            ip_address=ip_address,
            installed_at=timezone.now()
    )

        return Response(PluginInstallUninstallSerializer(plugin_action).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='uninstall')
    def uninstall(self, request):
        plugin_id = request.data.get('pluginId')

        try:
        # Find the plugin action and update uninstalled_at
           plugin_actions = PluginInstallUninstall.objects.filter(plugin_id=plugin_id, uninstalled_at__isnull=True)

           if not plugin_actions:
               return Response({'error': 'Plugin not found or already uninstalled'}, status=status.HTTP_404_NOT_FOUND)
           
            
           for action in plugin_actions:
               action.uninstalled_at = timezone.now()
               action.save()
            
           serializer = PluginInstallUninstallSerializer(plugin_actions, many=True)
           return Response(serializer.data, status=status.HTTP_200_OK)
        except PluginInstallUninstall.DoesNotExist:
           return Response({'error': 'Plugin not found or already uninstalled'}, status=status.HTTP_404_NOT_FOUND)
       
       

class PluginEnableDisableViewSet(viewsets.ViewSet):
    """
    A ViewSet for handling plugin enabling and disabling actions.
    """

    @action(detail=False, methods=['post'], url_path='enable')
    def enable(self, request):
        plugin_id = request.data.get('pluginId')

        try:
            # Get the most recent plugin action where the plugin is installed and not yet uninstalled
            plugin_action = PluginInstallUninstall.objects.filter(plugin_id=plugin_id, uninstalled_at__isnull=True).latest('installed_at')
            # Check if the plugin is already enabled
            if PluginEnableDisable.objects.filter(plugin_install_uninstall=plugin_action, enabled_at__isnull=False, disabled_at__isnull=True).exists():
                return Response({'error': 'The plugin is already enabled'}, status=status.HTTP_400_BAD_REQUEST)

            # Create a new entry for plugin enabling
            enable_disable_action = PluginEnableDisable.objects.create(
                plugin_install_uninstall=plugin_action,
                enabled_at=timezone.now()
            )

            return Response(PluginEnableDisableSerializer(enable_disable_action).data, status=status.HTTP_201_CREATED)
        
        except PluginInstallUninstall.DoesNotExist:
            return Response({'error': 'No active installation found for this plugin'}, status=status.HTTP_404_NOT_FOUND)
        
        
    @action(detail=False, methods=['post'], url_path='disable')
    def disable(self, request):
        plugin_id = request.data.get('pluginId')

        try:
        # Get the most recent plugin action where the plugin is installed and not yet uninstalled
            plugin_action = PluginInstallUninstall.objects.filter(plugin_id=plugin_id, uninstalled_at__isnull=True).latest('installed_at')

        # Check if the plugin is enabled (enabled_at is not null and disabled_at is null)
            enable_disable_action = PluginEnableDisable.objects.filter(
                plugin_install_uninstall=plugin_action, 
                enabled_at__isnull=False, 
                disabled_at__isnull=True
                ).first()

            if enable_disable_action:
            # If the plugin is currently enabled, update the `disabled_at` field to mark it as disabled
                enable_disable_action.disabled_at = timezone.now()
                enable_disable_action.save()

                return Response(PluginEnableDisableSerializer(enable_disable_action).data, status=status.HTTP_200_OK)
            else:
            # If the plugin is not enabled or already disabled, return an error
                return Response({'error': 'The plugin is either already disabled or was never enabled'}, status=status.HTTP_400_BAD_REQUEST)

        except PluginInstallUninstall.DoesNotExist:
            return Response({'error': 'No active installation found for this plugin'}, status=status.HTTP_404_NOT_FOUND)

   
   
class PluginRegistrationCheckViewSet(viewsets.ViewSet):
    """
    A ViewSet for handling plugin registration and license validation.
    """

    @action(detail=False, methods=['post'], url_path='check-registration')
    def check_registration(self, request):
        """
        Check if a plugin is registered and validates the associated license.
        """
        plugin_id = request.data.get('pluginId')

        if not plugin_id:
            return Response({'error': 'Plugin ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Step 1: Retrieve the plugin using plugin_id
            plugin = PluginMaster.objects.get(plugin_id=plugin_id)

            # Step 2: Check if the plugin has an associated license
            license = plugin.license_id

            # Step 3: Validate if the license is active and valid based on date
            now = timezone.now()
            if license.valid_from <= now <= license.valid_till:
                return Response({
                    'message': 'Plugin is registered and the license is valid.',
                    # 'plugin': PluginSerializer(plugin).data,
                    # 'license': LicenseSerializer(license).data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'error': 'License has expired or is not yet valid.'
                }, status=status.HTTP_403_FORBIDDEN)
        
        except PluginMaster.DoesNotExist:
            return Response({
                'error': 'Plugin not found.'
            }, status=status.HTTP_404_NOT_FOUND)
        except License.DoesNotExist:
            return Response({
                'error': 'License not found for this plugin.'
            }, status=status.HTTP_404_NOT_FOUND)


@csrf_exempt
@require_http_methods(["POST"])
def cdr_resposne_to_ai(request):
   
    if request.method != 'POST':
        return JsonResponse({"error": "Invalid request method"}, status=405)

    
    try:
       
        id = request.POST.get('id')
        msg_id = request.POST.get('msg_id')
        status = request.POST.get('status')
        cdr_file = request.FILES.get('cdr_file')  

        
        if not msg_id or not status:
            return JsonResponse({"error": "All fields (id, msg_id, status, cdr_file) are required"}, status=400)

        time.sleep(1)
        
        email_detail, created = EmailDetails.objects.update_or_create(
            id=id,
            defaults={
                'status': status,  
                'cdr_file': cdr_file,
            }
        )
        return JsonResponse({
            'status': 200,
            'message': 'Status Updated successfully.',
            'data':{'id': email_detail.id},
            'errors': None
        }, status=200)

    except Exception as e:
        return JsonResponse({"error": "An unexpected error occurred", "details": str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def url_response_to_ai(request):
    try:
        data = json.loads(request.body)
        id = data.get('id')
        msg_id = data.get('msg_id')
        status = data.get('status')
        url = data.get('url')

        if not id or not msg_id or not status or not url:
            return JsonResponse({"error": "All fields (id, msg_id, status, url) are required"}, status=400)
    
        time.sleep(1)
        email_detail, created = EmailDetails.objects.update_or_create(
            id=id,
            defaults={
                'status': status,
            }
        )
        return JsonResponse({
            'status': 200,
            'message': 'Status updated successfully.',
            'data': {'id': email_detail.id},
            'errors': None
        }, status=200)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format"}, status=400)
    except Exception as e:
        return JsonResponse({"error": "An unexpected error occurred", "details": str(e)}, status=500)
    

@csrf_exempt
@require_http_methods(["POST"])
def content_response_to_ai(request):
    try:
        
        data = json.loads(request.body)
        id = data.get('id')
        msg_id = data.get('msg_id')
        from_email = data.get('from_email')
        to_email = data.get('to_email')
        status = data.get('status')
        content = data.get('content')

       
        if not from_email or not msg_id or not status or not to_email:
            return JsonResponse({"error": "All fields ( msg_id, status,from_email,to_email ) are required"}, status=400)

        time.sleep(1)

        email_detail, created = EmailDetails.objects.update_or_create(
            id=id,
            defaults={
                'status': status,
            }
        )
        return JsonResponse({
            'status': 200,
            'message': 'Status updated successfully.',
            'data': {'id': email_detail.id},
            'errors': None
        }, status=200)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format"}, status=400)

    except Exception as e:
        return JsonResponse({"error": "An unexpected error occurred", "details": str(e)}, status=500)
    

class BrowserDetailsViewSet(viewsets.ModelViewSet):
    queryset = LicenseAllocation.objects.all()
    serializer_class = LicenseAllocationSerializer

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        data = {
            'license_validity': instance.license_validity,
            'license_allocated_to': instance.license_allocated_to,
            'allocated_date': instance.allocated_date,
        }
        print(
            f"License Validity: {instance.license_validity}",
            f"License Allocated To: {instance.license_allocated_to}",
            f"Allocated Date: {instance.allocated_date}",
        )

        return Response(data)

class GetDisputesView(viewsets.ViewSet):
    queryset = Dispute.objects.all()
    serializer_class = DisputeSerializer
    def list(self, request):
        try:
            disputes = Dispute.objects.all().order_by('-created_at')
            dispute_data = []
            for dispute in disputes:
                dispute_info = DisputeInfo.objects.filter(dispute=dispute).first()
                dispute_details = {
                    'dispute_id': dispute.id,
                    'email': dispute.email,
                    'msg_id': dispute.msg_id,
                    'counter': dispute.counter,
                    'status': dispute.status,
                    'created_at': dispute.created_at,
                    'user_comment': dispute_info.user_comment if dispute_info else None
                }
                dispute_data.append(dispute_details)
            return JsonResponse({
                    "message": "Disputes retrieved successfully",
                    "STATUS": "Success",
                    "Code": 1,
                    "data": dispute_data
                }, safe=False)
        except Exception as e:
            return JsonResponse({
                "message": str(e),
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=500)
    

@csrf_exempt
def spam_email(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email_id = data.get('emailId')
            plugin_id = data.get('pluginId') 
            
            if not email_id:
                return JsonResponse({
                    "message": "email_id is missing",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=400)

            print(f"Looking for emails with receiver: {email_id}")

            # Query using Django ORM
            spam_emails = EmailDetails.objects.filter(
                status__in=['unsafe','pending'],
                recievers_email=email_id
            ).order_by('-create_time')

            print(spam_emails)
            
            serialized_data = serialize('json', spam_emails)
            
            
            parsed_data = json.loads(serialized_data)

            # Extract the actual data from the serialized format
            result = [item['fields'] for item in parsed_data]

            return JsonResponse(result, safe=False)

        except json.JSONDecodeError:
            return JsonResponse({
                "message": "Invalid JSON format",
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=400)

    return JsonResponse({
        "message": "Invalid request method",
        "STATUS": "Error",
        "Code": 0,
        "data": ""
    }, status=405)

@csrf_exempt
def graph_count(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email_id = data.get('emailId')

            if not email_id:
                return JsonResponse({
                    "message": "email_id is missing",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=400)

            total_disputes = Dispute.objects.filter(email=email_id).count()
            total_processed_emails = EmailDetails.objects.filter(recievers_email=email_id).count()
            total_spam_emails = EmailDetails.objects.filter(status='unsafe', recievers_email=email_id).count()

            result_data = {
                'total_disputes': total_disputes,
                'total_processed_emails': total_processed_emails,
                'total_spam_emails': total_spam_emails,
            }

            return JsonResponse({
                "message": "Counts retrieved successfully",
                "STATUS": "Success",
                "Code": 1,
                "data": result_data
            })

        except json.JSONDecodeError:
            return JsonResponse({
                "message": "Invalid JSON format",
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=400)
        except Exception as e:
            return JsonResponse({
                "message": str(e),
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=500)

    return JsonResponse({
        "message": "Invalid request method",
        "STATUS": "Error",
        "Code": 0,
        "data": ""
    }, status=405)


# for cdr attachments urls block 
url_pattern = re.compile(r'https?://\S+')

def replace_urls_in_text(text):
    return url_pattern.sub(lambda match: match.group().replace("http", "http[dot]").replace(".", "[dot]"), text)
def process_pdf(file_path):
    output_pdf_path = file_path.replace(".pdf", "_modified.pdf")
    reader = PdfReader(file_path)
    writer = PdfWriter()

    for page_num in range(len(reader.pages)):
        page = reader.pages[page_num]
        if '/Annots' in page:
            annots = page['/Annots']
            for annot in annots:
                annot_obj = annot.get_object()
                if annot_obj.get('/Subtype') == '/Link':
                   
                    if '/A' in annot_obj:
                        del annot_obj['/A']
                    if '/URI' in annot_obj:
                        del annot_obj['/URI']
        
        text = page.extract_text()
        if text:
            modified_text = replace_urls_in_text(text)
        writer.add_page(page)

    with open(output_pdf_path, "wb") as output_pdf_file:
        writer.write(output_pdf_file)

    return output_pdf_path

def process_docx(file_path):
    doc = docx.Document(file_path)
    for para in doc.paragraphs:
        para.text = replace_urls_in_text(para.text)

    modified_docx_path = file_path.replace(".docx", "_modified.docx")
    doc.save(modified_docx_path)
    return modified_docx_path


def process_excel(file_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        worksheet = wb[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = replace_urls_in_text(cell.value)

    modified_excel_path = file_path.replace(".xlsx", "_modified.xlsx")
    wb.save(modified_excel_path)
    return modified_excel_path
    

@csrf_exempt
def block_file_and_urls(request):
    if request.method == 'POST':
        try:
            
            if 'file' not in request.FILES:
                return JsonResponse({
                    "message": "No file provided",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=400)

            uploaded_file = request.FILES['file']
    
            file_name = default_storage.save(uploaded_file.name, ContentFile(uploaded_file.read()))
            file_path = default_storage.path(file_name)
            modified_file_path = ""
            if uploaded_file.name.endswith('.pdf'):
                modified_file_path = process_pdf(file_path)
            elif uploaded_file.name.endswith('.docx'):
                modified_file_path = process_docx(file_path)
            elif uploaded_file.name.endswith('.xlsx'):
                modified_file_path = process_excel(file_path)
            elif uploaded_file.name.endswith('.txt'):
                with open(file_path, 'r') as file:
                    file_content = file.read()
                    modified_content = replace_urls_in_text(file_content)
                modified_file_path = file_path.replace(".txt", "_modified.txt")
                with open(modified_file_path, 'w') as modified_file:
                    modified_file.write(modified_content)
            else:
                return JsonResponse({
                    "message": "Unsupported file type",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=400)
            os.remove(file_path)
            response_data = {
                "message": "File urls blocked successfully",
                "STATUS": "Success",
                "Code": 1,
                "data": {
                    "modified_file": request.build_absolute_uri(default_storage.url(modified_file_path)),
                    "file_name": os.path.basename(modified_file_path),
                    "file_size": os.path.getsize(modified_file_path)
                }
            }    
            return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({
                "message": str(e),
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=500)

    return JsonResponse({
        "message": "Invalid request method",
        "STATUS": "Error",
        "Code": 0,
        "data": ""
    }, status=405)



@csrf_exempt
def get_disputes_raise_data(request):
    if request.method == 'POST':
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)
            email = data.get('emailId')
            
            # Check if email is provided
            if not email:
                return JsonResponse({
                    "message": "Missing email",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=400)

            # Retrieve disputes based on the email
            disputes = Dispute.objects.filter(email=email)

            # Check if no disputes are found
            if not disputes.exists():
                return JsonResponse({
                    "message": "Dispute not found",
                    "STATUS": "Not found",
                    "Code": 0,
                    "data": 0,
                }, status=200)

            # Prepare response data for multiple disputes
            response_data = [{
                'dispute_id': dispute.id,
                'email': dispute.email,
                'msg_id': dispute.msg_id,
                'counter': dispute.counter,
                'status': dispute.status,
                'created_at': dispute.created_at,
            } for dispute in disputes]

            # Return the response with the dispute details
            return JsonResponse({
                "message": "Disputes retrieved successfully",
                "STATUS": "Success",
                "Code": 1,
                "data": response_data
            })

        except Exception as e:
            # Handle unexpected errors
            return JsonResponse({
                "message": str(e),
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=500)
@csrf_exempt
def get_allocation_data(request):
    if request.method == 'POST':
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)
            hashed_license_id = data.get('licenseId')

            # Check if license_id is provided
            if not hashed_license_id:
                return JsonResponse({
                    "message": "Missing licenseId",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=400)

            # Retrieve the License record using the hashed_license_id
            license = License.objects.filter(hashed_license_id=hashed_license_id).first()

            if not license:
                return JsonResponse({
                    "message": "License not found",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=404)

            # Prepare the response data
            response_data = {
                'allocated_to': license.allocated_to,
                'valid_from': license.valid_from.strftime('%Y-%m-%d %H:%M:%S') if license.valid_from else "N/A",
                'valid_till': license.valid_till.strftime('%Y-%m-%d %H:%M:%S') if license.valid_till else "N/A"
            }

            # Return the response with allocation data
            return JsonResponse({
                "message": "License allocation data retrieved successfully",
                "STATUS": "found",
                "Code": 1,
                "data": response_data
            })

        except json.JSONDecodeError:
            return JsonResponse({
                "message": "Invalid JSON format",
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=400)

        except Exception as e:
            return JsonResponse({
                "message": f"Internal Server Error: {str(e)}",
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=500)
        
@csrf_exempt
def get_counter_count(request):
    if request.method == 'POST':
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)
            msg_id = data.get('messageId')

            if not msg_id:
                return JsonResponse({
                    "message": "Missing messageId",
                    "STATUS": "Error",
                    "Code": 0,
                    "data": ""
                }, status=400)

            # Fetch the existing Dispute object based on messageId
            dispute = Dispute.objects.filter(msg_id=msg_id).first()
            if not dispute:
                return JsonResponse({
                    "message": "Dispute entry not found",
                    "STATUS": "Not Found",
                    "Code": 0,
                    "counter": 0
                }, status=200)

            # Return the existing counter value
            return JsonResponse({
                "message": "Counter count retrieved successfully",
                "STATUS": "Found",
                "Code": 1,
                "counter": dispute.counter if dispute.counter is not None else 0
            })

        except json.JSONDecodeError:
            return JsonResponse({
                "message": "Invalid JSON format",
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=400)
        except Exception as e:
            return JsonResponse({
                "message": f"Internal Server Error: {str(e)}",
                "STATUS": "Error",
                "Code": 0,
                "data": ""
            }, status=500)
        
# 
class UpdateEmailDetailsView(APIView):

    def post(self, request):
        # Pass the request data to the serializer
        serializer = DisputeUpdateInfoSerializer(data=request.data)

        if serializer.is_valid():
            # Call the fetch method to retrieve the details
            email_detail, admin_comment = serializer.fetch(serializer.validated_data)
            return Response({
                "message": "Details fetched successfully",
                "data": {
                    "msg_id": email_detail.msg_id,
                    "status": email_detail.status,
                    "admin_remarks": admin_comment
                }
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@csrf_exempt
def browser_uninstall(request):
    """
    Handles browser unregistration for a user's system based on email and browser name.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            browser_name = data.get('browser')

            if not email or not browser_name:
                return JsonResponse({"error": "email and browser are required fields."}, status=400)

            try:
                license_allocations = LicenseAllocation.objects.filter(allocated_to__iexact=email).order_by('-allocation_date')
            except Exception as e:
                return JsonResponse({"error": f"Error fetching license allocation: {str(e)}"}, status=500)

            if not license_allocations.exists():
                return JsonResponse({"error": "No license allocation found for the provided email."}, status=404)

            license_allocation = license_allocations.first()

            try:
                user_system = UserSystemDetails.objects.filter(license_allocation=license_allocation).first()
            except UserSystemDetails.DoesNotExist:
                return JsonResponse({"error": "User system not found for the given license."}, status=404)

            try:
                browser_details = SystemBrowserDetails.objects.get(
                    device_information=user_system, 
                    browser=browser_name
                )
                
                # Update both unregistered_at and is_active
                browser_details.unregistered_at = now()
                browser_details.is_active = False
                browser_details.save()
                
                return JsonResponse({
                    "success": f"Browser '{browser_name}' unregistered successfully.",
                    "is_active": False
                }, status=200)

            except SystemBrowserDetails.DoesNotExist:
                return JsonResponse({"error": "Browser not found for the specified license and system."}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data."}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}"}, status=500)

    return JsonResponse({"error": "Invalid request method. Only POST is allowed."}, status=405)

def extract_email(email_string):
    """Extract clean email address from different formats"""
    import re
    pattern = r'(?:"?([^"]*)"?\s)?(?:<)?([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})(?:>)?'
    match = re.search(pattern, email_string)
    return match.group(2) if match else email_string

@csrf_exempt
def raise_dispute_view(request):
    """
    Handle raising a dispute and creating the related dispute info in a single function.
    Prevent raising a dispute if the counter in Dispute is >= 3.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_id = request.user.id
            
            # Clean the email address from the input
            clean_email = extract_email(data["email"])

            # Fetch EmailDetails with cleaned email
            email_details_qs = EmailDetails.objects.filter(
                recievers_email__icontains=clean_email, 
                msg_id=data["msgId"]
            )

            if not email_details_qs.exists():
                return JsonResponse({"error": "Email details not found."}, status=404)

            email_details = email_details_qs.first()

            # Use cleaned email for dispute lookup
            dispute_qs = Dispute.objects.filter(email=clean_email, msg_id=data["msgId"])
            
            if dispute_qs.exists():
                dispute = dispute_qs.first()

                if dispute.counter >= 3:
                    return JsonResponse(
                        {"error": "Cannot raise dispute. Counter has reached the limit of 3."},
                        status=400
                    )

                dispute.counter = dispute.counter + 1 if dispute.counter else 1
                dispute.updated_by_id = user_id
                dispute.save()
            else:
                dispute_data = {
                    'email': clean_email,
                    'msg_id': data["msgId"],
                    'counter': 1,
                    'status': Dispute.UNSAFE,
                    'created_by_id': user_id,
                    'updated_by_id': user_id,
                    'emaildetails': email_details,
                }
                dispute = Dispute.objects.create(**dispute_data)

            dispute_info_data = {
                'dispute': dispute,
                'user_comment': data["userComment"],
                'counter': dispute.counter,
                'created_by_id': user_id,
                'updated_by_id': user_id,
            }
            dispute_info = DisputeInfo.objects.create(**dispute_info_data)

            response_data = {
                "dispute_id": dispute.id,
                "email": dispute.email,
                "msg_id": dispute.msg_id,
                "status": dispute.get_status_display(),
                "counter": dispute.counter,
                "user_comment": dispute_info.user_comment,
                "dispute_info_id": dispute_info.id,
            }

            return JsonResponse(response_data, status=201)

        except EmailDetails.DoesNotExist:
            return JsonResponse({"error": "Email details not found."}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Invalid HTTP method. Use POST."}, status=405)
@csrf_exempt
@api_view(['POST'])
def pending_status_check(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            msg_id = data.get('msgId')
            email = data.get('email')

            
            email_details = EmailDetails.objects.filter(
                msg_id=msg_id,
                status='pending'
            ).first()

            if not email_details:
                return JsonResponse({"error": "No pending email found for given message ID."}, status=404)

            return Response({
                "status": email_details.status,
                "msgId": email_details.msg_id
            }, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format."}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An error occurred: {str(e)}"}, status=500)

    return JsonResponse({"error": "Invalid request method. Only POST is allowed."}, status=405)